"""Exportadores de relatórios — Anexo I 2.9.2 (CSV / Excel / PDF). Tudo open-source."""
import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .reports import Relatorio

ACCENT = colors.HexColor("#4f46e5")


def to_csv(rel: Relatorio) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")  # ';' abre direto no Excel pt-BR
    writer.writerow(rel.colunas)
    writer.writerows(rel.linhas)
    return buf.getvalue().encode("utf-8-sig")  # BOM p/ acentuação no Excel


def to_xlsx(rel: Relatorio) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"
    ws.append(rel.colunas)
    header_fill = PatternFill("solid", fgColor="4F46E5")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    for row in rel.linhas:
        ws.append([str(c) for c in row])
    for col in ws.columns:
        largura = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(largura + 2, 12), 60)
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_pdf(rel: Relatorio) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        topMargin=1.4 * cm, bottomMargin=1.2 * cm, leftMargin=1.2 * cm, rightMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=7.5, leading=9.5)
    head_style = ParagraphStyle("head", parent=cell_style, textColor=colors.white, fontName="Helvetica-Bold")

    ncols = len(rel.colunas)
    largura_total = landscape(A4)[0] - 2.4 * cm
    col_widths = [largura_total / ncols] * ncols

    data = [[Paragraph(str(c), head_style) for c in rel.colunas]]
    for row in rel.linhas:
        data.append([Paragraph(str(c), cell_style) for c in row])

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), ACCENT),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d6dae1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f7f9")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ])
    )

    elems = [
        Paragraph(rel.titulo, styles["Title"]),
        Paragraph("ThermoProActive — Gestão de Manutenção Preditiva", styles["Italic"]),
        Spacer(1, 0.4 * cm),
        table,
    ]
    doc.build(elems)
    return buf.getvalue()


EXPORTERS = {"csv": to_csv, "xlsx": to_xlsx, "pdf": to_pdf}
CONTENT_TYPES = {
    "csv": "text/csv; charset=utf-8",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}
